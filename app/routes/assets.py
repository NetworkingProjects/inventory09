
from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file, jsonify, session, current_app
from flask_login import login_required, current_user
from sqlalchemy import or_
from ..extensions import db
from ..models import Asset, Team, Manufacturer, VendorM, LocationM, Recipient, CategoryM, SubCategoryM
from ..forms import AssetForm
import io, csv, os, uuid
import pandas as pd
from openpyxl import load_workbook
from datetime import date, timedelta, datetime
from werkzeug.utils import secure_filename





DATE_HEADERS = {"invoice_date", "received_date", "last_calibrated", "next_calibration"}

def _cell_to_json(value, header_name):
    """Normalize Excel cell to JSON-safe value for session storage."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    # allow Excel serials; we'll convert in parse_date() at commit time
    if header_name in DATE_HEADERS and isinstance(value, (int, float)):
        return value
    return str(value).strip()




assets_bp = Blueprint("assets", __name__)

def can_create(): return current_user.is_authenticated and current_user.has("create")
def can_export(): return current_user.is_authenticated and current_user.has("export")
def can_delete(): return current_user.is_authenticated and (getattr(current_user, "role", None) in ("superadmin", "admin"))


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")

def parse_date(value):
    """
    Accepts: None, "", date, datetime, Excel serial (int/float), or string dates like:
    YYYY-MM-DD, DD-MM-YYYY, YYYY/MM/DD, DD/MM/YYYY, MM/DD/YYYY, MM-DD-YYYY.
    Returns: datetime.date or None.
    """
    if _is_blank(value):
        return None
    if isinstance(value, date) and not isinstance(value, date):
        return value
    if isinstance(value, date):
        return value.date()

    # Excel serial (xlsx)
    if isinstance(value, (int, float)):
        try:
            base = date(1899, 12, 30)  # Excel epoch
            serial = int(value)
            if serial > 0:
                return base + timedelta(days=serial)
        except Exception:
            pass

    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    try:
        from dateutil import parser as dateparser  # optional
        return dateparser.parse(s, dayfirst=False).date()
    except Exception:
        return None

def norm_ynna(value):
    """
    Normalize y/n/na (accepts yes/no/true/false/1/0 too).
    Returns: 'y', 'n', 'na' or None (if blank).
    """
    if _is_blank(value):
        return None
    s = str(value).strip().lower()
    if s in ("y", "yes", "true", "1"):
        return "y"
    if s in ("n", "no", "false", "0"):
        return "n"
    if s in ("na", "n/a"):
        return "na"
    return "na"

def clean_str(value):
    """Return stripped string or None if blank."""
    if _is_blank(value):
        return None
    return str(value).strip()



@assets_bp.route("/")
@login_required
def dashboard():
    q = request.args.get("q","").strip()
    category = request.args.get("category","").strip()
    location = request.args.get("location","").strip()
    query = Asset.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Asset.invoice_no.like(like), Asset.serial_number.like(like), Asset.model.like(like), Asset.description.like(like)))
    if category:
        query = query.filter(Asset.category==category)
    if location:
        query = query.filter(Asset.location==location)
    page = request.args.get("page", 1, type=int)
    assets = query.order_by(Asset.id.desc()).paginate(page=page, per_page=10)
    categories = [c.name for c in CategoryM.query.order_by(CategoryM.name.asc()).all()]
    locations = [l.name for l in LocationM.query.order_by(LocationM.name.asc()).all()]
    return render_template("assets/index.html", assets=assets, q=q, categories=categories, locations=locations, can_create=can_create(), can_export=can_export(), can_delete=can_delete(), )

def _set_choices(form: AssetForm):
    form.team.choices = [("", "")] + [(t.name, t.name) for t in Team.query.order_by(Team.name.asc()).all()]
    form.manufacturer.choices = [("", "")] + [(m.name, m.name) for m in Manufacturer.query.order_by(Manufacturer.name.asc()).all()]
    form.vendor.choices = [("", "")] + [(v.name, v.name) for v in VendorM.query.order_by(VendorM.name.asc()).all()]
    form.location.choices = [("", "")] + [(l.name, l.name) for l in LocationM.query.order_by(LocationM.name.asc()).all()]
    form.category.choices = [("", "")] + [(c.name, c.name) for c in CategoryM.query.order_by(CategoryM.name.asc()).all()]
    form.sub_category.choices = [("", "")] + [(s.name, s.name) for s in SubCategoryM.query.order_by(SubCategoryM.name.asc()).all()]

@assets_bp.route("/assets/create", methods=["GET","POST"])
@login_required
def create():
    if not can_create():
        flash("Permission denied", "error"); return redirect(url_for("assets.dashboard"))
    form = AssetForm(); _set_choices(form)
    recips = Recipient.query.order_by(Recipient.name.asc()).all()
    cats = CategoryM.query.order_by(CategoryM.name.asc()).all()
    if form.validate_on_submit():
        a = Asset(
            invoice_no=form.invoice_no.data, invoice_date=form.invoice_date.data, serial_number=form.serial_number.data,
            purchase_order_no=form.purchase_order_no.data, received_date=form.received_date.data, owner_email=form.owner_email.data,
            description=form.description.data, manufacturer=form.manufacturer.data, model=form.model.data, vendor=form.vendor.data,
            mfg_country=form.mfg_country.data, hsn_code=form.hsn_code.data, is_bonded=form.is_bonded.data, last_calibrated=form.last_calibrated.data,
            next_calibration=form.next_calibration.data, notes=form.notes.data, entry_no=form.entry_no.data, returnable_no=form.returnable_no.data,
            cap_x=form.cap_x.data, amortization_period=form.amortization_period.data, team=form.team.data,
            recipient_name=form.recipient_name.data, recipient_email=form.recipient_email.data, category=form.category.data,
            sub_category=form.sub_category.data, location=form.location.data
        )
        db.session.add(a); db.session.commit(); flash("Asset added", "success")
        return redirect(url_for("assets.dashboard"))
    return render_template("assets/form.html", form=form, mode="create", recips=recips, cats=cats)

@assets_bp.route("/assets/<int:id>/edit", methods=["GET","POST"])
@login_required
def edit(id):
    a = Asset.query.get_or_404(id)
    form = AssetForm(obj=a); _set_choices(form)
    recips = Recipient.query.order_by(Recipient.name.asc()).all()
    cats = CategoryM.query.order_by(CategoryM.name.asc()).all()
    if form.validate_on_submit():
        form.populate_obj(a); db.session.commit(); flash("Asset updated", "success")
        return redirect(url_for("assets.view", id=a.id))
    return render_template("assets/form.html", form=form, mode="edit", a=a, recips=recips, cats=cats)

@assets_bp.post("/assets/<int:id>/delete")
@login_required
def delete(id):
    if not can_delete():
        flash("Permission denied", "error")
        return redirect(url_for("assets.dashboard"))
    a = Asset.query.get_or_404(id)
    db.session.delete(a)
    db.session.commit()
    flash("Asset deleted", "success")
    return redirect(url_for("assets.dashboard"))


@assets_bp.route("/assets/<int:id>")
@login_required
def view(id):
    a = Asset.query.get_or_404(id)
    return render_template("assets/view.html", a=a)

@assets_bp.route("/export/excel")
@login_required
def export_excel():
    if not can_export(): flash("Permission denied", "error"); return redirect(url_for("assets.dashboard"))
    df = pd.read_sql(Asset.query.statement, db.session.bind)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Assets")
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="assets.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@assets_bp.route("/export/csv")
@login_required
def export_csv():
    if not can_export(): flash("Permission denied", "error"); return redirect(url_for("assets.dashboard"))
    si = io.StringIO(); cw = csv.writer(si)
    cols = [c.name for c in Asset.__table__.columns]; cw.writerow(cols)
    for a in Asset.query.order_by(Asset.id.desc()).all():
        cw.writerow([getattr(a, c) for c in cols])
    mem = io.BytesIO(si.getvalue().encode("utf-8"))
    return send_file(mem, as_attachment=True, download_name="assets.csv", mimetype="text/csv")

@assets_bp.route("/analytics")
@login_required
def analytics_page():
    return render_template("assets/analytics.html")

@assets_bp.route("/analytics.json")
@login_required
def analytics_json():
    from sqlalchemy import func
    by_category = db.session.query(Asset.category, func.count(Asset.id)).group_by(Asset.category).all()
    by_location = db.session.query(Asset.location, func.count(Asset.id)).group_by(Asset.location).all()
    today = date.today()
    due_soon = Asset.query.filter(Asset.next_calibration != None, Asset.next_calibration <= today + timedelta(days=30)).count()
    ok = Asset.query.filter(Asset.next_calibration != None, Asset.next_calibration > today + timedelta(days=30)).count()
    return jsonify({
        "by_category":[{"label":k or "Unknown","value":v} for k,v in by_category],
        "by_location":[{"label":k or "Unknown","value":v} for k,v in by_location],
        "calibration":[{"label":"Due ≤30d","value":due_soon},{"label":"OK","value":ok}]
    })

# Import
EXPECTED_COLS = ["invoice_no","invoice_date","serial_number","purchase_order_no","received_date","owner_email","description","manufacturer","model","vendor","mfg_country","hsn_code","is_bonded","last_calibrated","next_calibration","notes","entry_no","returnable_no","cap_x","amortization_period","team","recipient_name","recipient_email","category","sub_category","location"]

def _upload_path(filename):
    os.makedirs(current_app.config["UPLOAD_FOLDER"], exist_ok=True)
    token = str(uuid.uuid4())
    return os.path.join(current_app.config["UPLOAD_FOLDER"], token + "_" + secure_filename(filename))

def _read_df(path):
    fn = path.lower()
    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        return pd.read_excel(path)
    return pd.read_csv(path)

@assets_bp.route("/import", methods=["GET"])
@login_required
def import_page():
    return render_template("assets/import.html")




# app/routes/assets.py
@assets_bp.route("/import/preview", methods=["POST"])
@login_required
def import_preview():
    if not can_create():
        flash("Permission denied", "error")
        return redirect(url_for("assets.dashboard"))
        
    f = request.files.get("file")
    if not f:
        flash("No file uploaded", "error")
        return redirect(url_for("assets.import_page"))

    filename = (f.filename or "").lower()
    required_headers = [
        "invoice_no","invoice_date","serial_number","purchase_order_no","received_date",
        "owner_email","description","manufacturer","model","vendor","mfg_country","hsn_code",
        "is_bonded","last_calibrated","next_calibration","notes","entry_no","returnable_no",
        "cap_x","amortization_period","team","recipient_name","recipient_email",
        "category","sub_category","location"
    ]

    rows = []
    errors = []
    warnings = []

    if filename.endswith(".csv"):
        try:
            from io import TextIOWrapper
            import csv
            wrapper = TextIOWrapper(f.stream, encoding="utf-8", newline="")
            reader = csv.DictReader(wrapper)
            if not reader.fieldnames:
                flash("CSV has no header row.", "error")
                return redirect(url_for("assets.import_page"))
            
            # Check for missing and extra headers
            file_headers = [h.strip() for h in reader.fieldnames if h]
            missing = [h for h in required_headers if h not in file_headers]
            extra = [h for h in file_headers if h not in required_headers]
            
            if missing:
                errors.append(f"Missing required headers: {', '.join(missing)}")
            if extra:
                warnings.append(f"Extra headers will be ignored: {', '.join(extra)}")
            
            # Read data rows
            for row_num, r in enumerate(reader, start=2):
                row_data = {}
                for k in required_headers:
                    value = r.get(k, "")
                    row_data[k] = value if value is not None else ""
                rows.append(row_data)
                
        except Exception as e:
            flash(f"Error reading CSV file: {str(e)}", "error")
            return redirect(url_for("assets.import_page"))

    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(f, data_only=True)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            file_headers = [str(v).strip() if v is not None else "" for v in header_row]
            file_headers = [h for h in file_headers if h]  # Remove empty headers

            # Check for missing and extra headers
            missing = [h for h in required_headers if h not in file_headers]
            extra = [h for h in file_headers if h not in required_headers]
            
            if missing:
                errors.append(f"Missing required headers: {', '.join(missing)}")
            if extra:
                warnings.append(f"Extra headers will be ignored: {', '.join(extra)}")

            idx_map = {name: i for i, name in enumerate(file_headers)}
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip completely empty rows
                    continue
                    
                row_data = {}
                for h in required_headers:
                    i = idx_map.get(h)
                    if i is not None and i < len(row):
                        row_data[h] = _cell_to_json(row[i], h)
                    else:
                        row_data[h] = ""
                rows.append(row_data)
                
        except Exception as e:
            flash(f"Error reading Excel file: {str(e)}", "error")
            return redirect(url_for("assets.import_page"))

    else:
        flash("Unsupported file type. Upload .csv or .xlsx", "error")
        return redirect(url_for("assets.import_page"))

    if not rows:
        flash("No data rows found in the file", "error")
        return redirect(url_for("assets.import_page"))

    # save raw rows to session (commit route will coerce types)
    session["import_rows"] = rows

    # Validation and analytics
    empty_counts = {h: 0 for h in required_headers}
    type_issues = []
    invalid_rows = 0
    
    for r in rows:
        for h in required_headers:
            if r.get(h, "") in ("", None):
                empty_counts[h] += 1
        
        # Check email format
        for email_field in ['owner_email', 'recipient_email']:
            email_val = r.get(email_field, "")
            if email_val and "@" not in email_val:
                type_issues.append({"column": email_field, "count": 1})

    # Check for date parse issues
    date_fields = ["invoice_date","received_date","last_calibrated","next_calibration"]
    bad_date_rows = 0
    for r in rows:
        for f in date_fields:
            v = r.get(f, "")
            if v not in ("", None):
                if parse_date(v) is None:
                    bad_date_rows += 1
                    break

    # Calculate invalid rows
    invalid_rows = bad_date_rows + len([t for t in type_issues if t["count"] > 0])

    stats = {
        "total_rows": len(rows),
        "invalid_rows": invalid_rows,
        "missing_cols": [h for h in required_headers if h not in (file_headers if 'file_headers' in locals() else [])],
        "extra_cols": extra if 'extra' in locals() else [],
        "empties": empty_counts,
        "type_issues": type_issues,
        "bad_date_rows": bad_date_rows
    }

    has_errors = bool(errors or bad_date_rows > 0 or invalid_rows > 0)

    return render_template(
        "assets/import_preview.html",
        rows_preview=rows[:25],
        total=len(rows),
        stats=stats,
        errors=errors,
        warnings=warnings,
        has_errors=has_errors
    )







@assets_bp.route("/import/commit", methods=["POST"])
@login_required
def import_commit():
    rows = session.get("import_rows") or []
    if not rows:
        flash("Nothing to import (no rows found). Please upload again.", "warning")
        return redirect(url_for("assets.import_page"))

    created = 0
    failed = 0
    fail_examples = []

    for idx, r in enumerate(rows, start=2):  # start=2 to reflect CSV line numbers
        try:
            asset = Asset(
                invoice_no           = clean_str(r.get("invoice_no")),
                invoice_date         = parse_date(r.get("invoice_date")),
                serial_number        = clean_str(r.get("serial_number")),
                purchase_order_no    = clean_str(r.get("purchase_order_no")),
                received_date        = parse_date(r.get("received_date")),
                owner_email          = clean_str(r.get("owner_email")),
                description          = clean_str(r.get("description")),
                manufacturer         = clean_str(r.get("manufacturer")),
                model                = clean_str(r.get("model")),
                vendor               = clean_str(r.get("vendor")),
                mfg_country          = clean_str(r.get("mfg_country")),
                hsn_code             = clean_str(r.get("hsn_code")),
                is_bonded            = norm_ynna(r.get("is_bonded")),
                last_calibrated      = parse_date(r.get("last_calibrated")),
                next_calibration     = parse_date(r.get("next_calibration")),
                notes                = clean_str(r.get("notes")),
                entry_no             = clean_str(r.get("entry_no")),
                returnable_no        = norm_ynna(r.get("returnable_no")),
                cap_x                = norm_ynna(r.get("cap_x")),
                amortization_period  = clean_str(r.get("amortization_period")),
                team                 = clean_str(r.get("team")),
                recipient_name       = clean_str(r.get("recipient_name")),
                recipient_email      = clean_str(r.get("recipient_email")),
                category             = clean_str(r.get("category")),
                sub_category         = clean_str(r.get("sub_category")),
                location             = clean_str(r.get("location")),
            )
            db.session.add(asset)
            db.session.flush()  # validate row-by-row
            created += 1
        except Exception as e:
            db.session.rollback()
            failed += 1
            if len(fail_examples) < 5:
                fail_examples.append(f"Row {idx}: {e}")

    db.session.commit()
    session.pop("import_rows", None)

    if failed == 0:
        flash(f"Imported {created} assets successfully.", "success")
    else:
        flash(f"Imported {created} assets, {failed} failed.", "warning")
        if fail_examples:
            flash("Examples: " + " | ".join(fail_examples), "warning")

    return redirect(url_for("assets.dashboard"))
